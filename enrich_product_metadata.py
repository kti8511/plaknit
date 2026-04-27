import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_FILE = ROOT / "data.json"
DESKTOP = Path.home() / "Desktop" / "26SS"


def latest_product_data_file():
    files = [p for p in DESKTOP.glob("*DATA.xlsx") if not p.name.startswith("~$")]
    if not files:
        raise FileNotFoundError("상품DATA.xlsx 파일을 찾지 못했습니다.")
    return max(files, key=lambda p: p.stat().st_mtime)


def load_product_meta(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["품번종합집계표"]
    header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    idx = {h: i for i, h in enumerate(header)}
    meta = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        sku = str(row[idx["품번"]] or "").strip() if idx.get("품번") is not None else ""
        if not sku:
            continue
        meta[sku] = {
            "season": str(row[idx["시즌"]] or "").strip(),
            "category_large": str(row[idx["복종(대)"]] or "").strip(),
            "category_small": str(row[idx["복종(소)"]] or "").strip(),
        }
    return meta


def main():
    product_file = latest_product_data_file()
    meta = load_product_meta(product_file)
    rows = json.loads(DATA_FILE.read_text(encoding="utf-8"))

    matched = 0
    for row in rows:
        sku = str(row.get("match_sku") or "").strip()
        info = meta.get(sku)
        if info:
            row.update(info)
            matched += 1
        else:
            row.setdefault("season", "")
            row.setdefault("category_large", "")
            row.setdefault("category_small", "")

    DATA_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"product_file": str(product_file), "meta_rows": len(meta), "matched_rows": matched, "data_rows": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
