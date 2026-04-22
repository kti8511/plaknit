import json
import csv
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\user\Documents\Codex\2026-04-20-files-mentioned-by-the-user-data")
WORKBOOK = ROOT / "상품DATA_관리본.xlsx"
JSON_DATA = ROOT / "selfmall_dashboard_data.json"
HTML_OUT = ROOT / "상품DATA_운영대시보드.html"
CSV_PATH = Path(r"C:\Users\user\Desktop\plaknit_20260420_477_16de.csv")
RETAILERS = ["자사몰", "무신사", "29cm", "글로리어스워커", "4XR", "애슬러", "롯데온"]


def money(value):
    return f"{int(value):,}"


def pct(value):
    return f"{value * 100:.1f}%"


def dec(value):
    return Decimal(str(value or "0").replace(",", "").strip() or "0")


def esc(value):
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def load_workbook_summary():
    wb = load_workbook(WORKBOOK, data_only=False)
    main = wb["품번종합집계표"]
    self_summary = wb["자사몰_요약"]
    self_agg = wb["자사몰_취합"]

    main_count = 0
    original_stock_value = 0
    original_order_qty = 0
    original_inventory_qty = 0
    stock_by_sku = {}
    for row in range(3, main.max_row + 1):
        sku = main.cell(row, 7).value
        if not sku:
            continue
        received_qty = main.cell(row, 18).value if isinstance(main.cell(row, 18).value, (int, float)) else 0
        sold_qty = main.cell(row, 21).value if isinstance(main.cell(row, 21).value, (int, float)) else 0
        stock_qty = main.cell(row, 25).value if isinstance(main.cell(row, 25).value, (int, float)) else 0
        main_count += 1
        original_order_qty += main.cell(row, 15).value if isinstance(main.cell(row, 15).value, (int, float)) else 0
        original_inventory_qty += stock_qty
        original_stock_value += main.cell(row, 26).value if isinstance(main.cell(row, 26).value, (int, float)) else 0
        stock_by_sku[str(sku)] = {
            "received_qty": received_qty,
            "main_sold_qty": sold_qty,
            "stock_qty": stock_qty,
        }

    summary = {}
    for row in range(2, self_summary.max_row + 1):
        key = self_summary.cell(row, 1).value
        value = self_summary.cell(row, 2).value
        if key:
            summary[str(key)] = value

    status_counts = Counter()
    unmatched_rows = []
    rows = []
    headers = [self_agg.cell(1, col).value for col in range(1, self_agg.max_column + 1)]
    for r in range(2, self_agg.max_row + 1):
        item = {headers[c - 1]: self_agg.cell(r, c).value for c in range(1, self_agg.max_column + 1)}
        if not item.get("상품번호"):
            continue
        status_counts[item.get("매칭상태") or ""] += 1
        rows.append(item)
        if item.get("매칭상태") == "미매칭":
            unmatched_rows.append(item)

    return {
        "main_count": main_count,
        "original_order_qty": original_order_qty,
        "original_inventory_qty": original_inventory_qty,
        "original_stock_value": original_stock_value,
        "self_summary": summary,
        "status_counts": dict(status_counts),
        "unmatched_rows": unmatched_rows,
        "agg_rows": rows,
        "stock_by_sku": stock_by_sku,
    }


def load_sales_calendar():
    daily = defaultdict(lambda: {"payment": 0, "gross": 0, "orders": 0, "qty": 0})
    retailer = defaultdict(lambda: {"payment": 0, "gross": 0, "orders": 0, "qty": 0})
    if not CSV_PATH.exists():
        return [], []

    orders = {}
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        next(reader, None)
        for row in reader:
            order_no = row[2]
            if not order_no:
                continue
            date = row[19][:10]
            mall = "자사몰"
            qty = int(dec(row[10]))
            gross = int(dec(row[10]) * dec(row[11]))
            payment = int(dec(row[6]))
            if order_no not in orders:
                orders[order_no] = {"date": date, "mall": mall, "payment": payment, "gross": 0, "qty": 0}
            orders[order_no]["gross"] += gross
            orders[order_no]["qty"] += qty

    for order in orders.values():
        date = order["date"]
        mall = order["mall"]
        daily[date]["payment"] += order["payment"]
        daily[date]["gross"] += order["gross"]
        daily[date]["orders"] += 1
        daily[date]["qty"] += order["qty"]
        retailer[mall]["payment"] += order["payment"]
        retailer[mall]["gross"] += order["gross"]
        retailer[mall]["orders"] += 1
        retailer[mall]["qty"] += order["qty"]

    daily_rows = [
        {
            "date": date,
            "payment": values["payment"],
            "gross": values["gross"],
            "orders": values["orders"],
            "qty": values["qty"],
            "avgOrder": round(values["payment"] / values["orders"]) if values["orders"] else 0,
        }
        for date, values in sorted(daily.items())
    ]
    retailer_rows = [
        {
            "retailer": name,
            "payment": values["payment"],
            "gross": values["gross"],
            "orders": values["orders"],
            "qty": values["qty"],
            "avgOrder": round(values["payment"] / values["orders"]) if values["orders"] else 0,
        }
        for name, values in sorted(retailer.items(), key=lambda item: item[1]["payment"], reverse=True)
    ]
    return daily_rows, retailer_rows


def make_dashboard():
    data = json.loads(JSON_DATA.read_text(encoding="utf-8"))
    wb_summary = load_workbook_summary()
    daily_rows, retailer_rows = load_sales_calendar()
    stock_by_sku = wb_summary["stock_by_sku"]

    for item in data:
        stock = stock_by_sku.get(str(item.get("match_sku") or ""), {})
        item["received_qty"] = stock.get("received_qty", 0)
        item["main_sold_qty"] = stock.get("main_sold_qty", 0)
        item["stock_qty"] = stock.get("stock_qty", 0)

    total_qty = sum(item.get("qty", 0) for item in data)
    total_payment = sum(item.get("payment", 0) for item in data)
    total_gross = sum(item.get("gross", 0) for item in data)
    matched = sum(1 for item in data if item.get("match_status") == "매칭됨")
    unmatched = sum(1 for item in data if item.get("match_status") == "미매칭")
    match_rate = matched / len(data) if data else 0
    avg_discount = 1 - (total_payment / total_gross) if total_gross else 0
    avg_unit = total_payment / total_qty if total_qty else 0

    top_payment = sorted(data, key=lambda item: item.get("payment", 0), reverse=True)[:12]
    all_dates = [
        datetime.fromisoformat(day["date"]).date()
        for item in data
        for day in item.get("daily", [])
        if day.get("date")
    ]
    latest_date = max(all_dates) if all_dates else None
    week_start = latest_date - timedelta(days=latest_date.weekday()) if latest_date else None
    week_end = week_start + timedelta(days=6) if week_start else None
    weekly_qty = defaultdict(int)
    for item in data:
        name = item.get("standard_name") or item.get("name") or ""
        for day in item.get("daily", []):
            if not day.get("date") or not week_start:
                continue
            sold_date = datetime.fromisoformat(day["date"]).date()
            if week_start <= sold_date <= week_end:
                weekly_qty[name] += int(day.get("qty") or 0)
    top_qty = sorted(
        [{"name": name, "qty": qty} for name, qty in weekly_qty.items() if qty],
        key=lambda item: item["qty"],
        reverse=True,
    )[:12]
    unmatched_rows = wb_summary["unmatched_rows"]

    by_product = defaultdict(lambda: {"qty": 0, "payment": 0, "gross": 0, "rows": 0, "received_qty": 0, "main_sold_qty": 0, "stock_qty": 0})
    by_type = defaultdict(lambda: {"qty": 0, "payment": 0})
    for item in data:
        key = item.get("standard_name") or item.get("name") or ""
        by_product[key]["qty"] += item.get("qty", 0)
        by_product[key]["payment"] += item.get("payment", 0)
        by_product[key]["gross"] += item.get("gross", 0)
        by_product[key]["rows"] += 1
        by_product[key]["received_qty"] += item.get("received_qty", 0) or 0
        by_product[key]["main_sold_qty"] += item.get("main_sold_qty", 0) or 0
        by_product[key]["stock_qty"] += item.get("stock_qty", 0) or 0
        by_product[key]["daily"] = by_product[key].get("daily", [])
        for day in item.get("daily", []):
            by_product[key]["daily"].append(day)
        by_type[item.get("source_type") or "단품"]["qty"] += item.get("qty", 0)
        by_type[item.get("source_type") or "단품"]["payment"] += item.get("payment", 0)

    product_rows = [
        {
            "name": name,
            "qty": values["qty"],
            "payment": values["payment"],
            "gross": values["gross"],
            "rows": values["rows"],
            "daily": values.get("daily", []),
            "received_qty": values["received_qty"],
            "main_sold_qty": values["main_sold_qty"],
            "stock_qty": values["stock_qty"],
            "discount": 1 - (values["payment"] / values["gross"]) if values["gross"] else 0,
            "salesRate": values["main_sold_qty"] / values["received_qty"] if values["received_qty"] else 0,
            "sellThrough": values["main_sold_qty"] / (values["main_sold_qty"] + values["stock_qty"]) if values["main_sold_qty"] + values["stock_qty"] else 0,
        }
        for name, values in by_product.items()
    ]
    product_rows.sort(key=lambda item: item["payment"], reverse=True)

    chart_labels = [item["date"] for item in daily_rows]
    chart_values = [item["payment"] for item in daily_rows]
    qty_labels = [item.get("name") or "" for item in top_qty[:8]]
    qty_values = [item.get("qty", 0) for item in top_qty[:8]]
    qty_week_label = f"{week_start:%Y-%m-%d} ~ {week_end:%Y-%m-%d}" if week_start and week_end else "주간"

    table_data = json.dumps(data, ensure_ascii=False)
    product_data = json.dumps(product_rows, ensure_ascii=False)
    unmatched_data = json.dumps(unmatched_rows, ensure_ascii=False, default=str)
    chart_data = json.dumps(
        {
            "paymentLabels": chart_labels,
            "paymentValues": chart_values,
            "qtyLabels": qty_labels,
            "qtyValues": qty_values,
            "qtyWeekLabel": qty_week_label,
        },
        ensure_ascii=False,
    )
    daily_data = json.dumps(daily_rows, ensure_ascii=False)
    retailer_data = json.dumps(retailer_rows, ensure_ascii=False)
    retailer_options = "\n".join(f'<option value="{esc(name)}">{esc(name)}</option>' for name in RETAILERS)

    max_payment = max([item.get("payment", 0) for item in top_payment] or [1])
    top_payment_rows = "\n".join(
        f"""
        <div class="rank-row">
          <div class="rank-name" title="{esc(item.get('standard_name') or item.get('name'))}">{esc(item.get('standard_name') or item.get('name'))}</div>
          <div class="rank-track"><div style="width:{(item.get('payment', 0) / max_payment) * 100:.1f}%"></div></div>
          <div class="rank-value">{money(item.get('payment', 0))}</div>
        </div>
        """
        for item in top_payment
    )

    unmatched_preview = "\n".join(
        f"""
        <tr>
          <td>{esc(row.get('상품번호'))}</td>
          <td>{esc(row.get('표준상품명'))}</td>
          <td>{esc(row.get('판매수량'))}</td>
          <td>{money(row.get('결제금액 배분액') or 0)}</td>
        </tr>
        """
        for row in unmatched_rows[:10]
    )

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>상품DATA 운영 대시보드</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <style>
    :root {{
      --bg: #f4f7fb;
      --panel: #ffffff;
      --ink: #182230;
      --muted: #667085;
      --line: #d7deea;
      --brand: #1f5f99;
      --teal: #0f766e;
      --amber: #b7791f;
      --red: #b42318;
      --green: #087443;
      --soft-blue: #e9f2fb;
      --soft-green: #e8f6f1;
      --soft-red: #fdeceb;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, 'Malgun Gothic', sans-serif; background: var(--bg); color: var(--ink); }}
    header {{ background: #fff; border-bottom: 1px solid var(--line); padding: 18px 24px; position: sticky; top: 0; z-index: 10; }}
    .header-inner {{ display: flex; align-items: center; justify-content: space-between; gap: 20px; }}
    .brand {{ display: flex; align-items: center; gap: 12px; }}
    .logo {{ width: 40px; height: 40px; border-radius: 8px; background: linear-gradient(135deg, var(--brand), var(--teal)); display: grid; place-items: center; color: #fff; font-weight: 800; }}
    h1 {{ margin: 0; font-size: 22px; letter-spacing: 0; }}
    .subtitle {{ color: var(--muted); font-size: 13px; margin-top: 3px; }}
    .status-pill {{ border: 1px solid var(--line); background: #fff; padding: 8px 10px; border-radius: 8px; font-size: 13px; color: var(--muted); white-space: nowrap; }}
    main {{ padding: 18px 24px 36px; }}
    .kpis {{ display: grid; grid-template-columns: repeat(6, minmax(130px, 1fr)); gap: 12px; margin-bottom: 16px; }}
    .kpi {{ background: var(--panel); border: 1px solid var(--line); border-radius: 8px; padding: 14px; min-height: 98px; }}
    .kpi .label {{ color: var(--muted); font-size: 12px; margin-bottom: 10px; }}
    .kpi .value {{ font-size: 23px; font-weight: 800; line-height: 1.15; }}
    .kpi .note {{ color: var(--muted); font-size: 12px; margin-top: 8px; }}
    .grid {{ display: grid; grid-template-columns: 1.2fr .8fr; gap: 16px; margin-bottom: 16px; }}
    .panel {{ background: var(--panel); border: 1px solid var(--line); border-radius: 8px; padding: 16px; min-width: 0; }}
    .panel h2 {{ margin: 0 0 12px; font-size: 17px; }}
    .charts {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
    .chart-box {{ min-height: 280px; }}
    canvas {{ width: 100% !important; height: 260px !important; }}
    .rank-row {{ display: grid; grid-template-columns: minmax(170px, 1fr) 1.2fr 90px; gap: 10px; align-items: center; padding: 7px 0; border-bottom: 1px solid #eef1f5; }}
    .rank-name {{ overflow: hidden; text-overflow: ellipsis; white-space: nowrap; font-size: 13px; }}
    .rank-track {{ height: 11px; background: #eef2f6; border-radius: 999px; overflow: hidden; }}
    .rank-track div {{ height: 100%; background: var(--brand); }}
    .rank-value {{ text-align: right; font-variant-numeric: tabular-nums; font-size: 13px; }}
    .toolbar {{ display: flex; gap: 10px; align-items: center; margin-bottom: 12px; flex-wrap: wrap; }}
    input, select {{ height: 38px; border: 1px solid var(--line); border-radius: 6px; background: #fff; color: var(--ink); padding: 0 10px; min-width: 190px; }}
    input[type="date"] {{ min-width: 150px; }}
    .tabs {{ display: flex; gap: 6px; margin: 0 0 16px; border-bottom: 1px solid var(--line); }}
    .tab-btn {{ height: 40px; border: 0; border-bottom: 3px solid transparent; background: transparent; padding: 0 14px; font-weight: 700; color: var(--muted); cursor: pointer; }}
    .tab-btn.active {{ color: var(--brand); border-bottom-color: var(--brand); }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .table-wrap {{ border: 1px solid var(--line); border-radius: 8px; overflow: auto; max-height: 560px; background: #fff; }}
    table {{ border-collapse: collapse; width: 100%; min-width: 1040px; font-size: 13px; }}
    th, td {{ padding: 9px 8px; border-bottom: 1px solid #edf0f4; text-align: left; vertical-align: top; }}
    th {{ position: sticky; top: 0; background: #edf3fb; z-index: 1; font-weight: 700; }}
    .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    .badge {{ display: inline-flex; align-items: center; height: 24px; padding: 0 8px; border-radius: 999px; font-size: 12px; border: 1px solid transparent; white-space: nowrap; }}
    .matched {{ background: var(--soft-green); color: var(--green); }}
    .unmatched {{ background: var(--soft-red); color: var(--red); }}
    .source {{ background: var(--soft-blue); color: var(--brand); }}
    .summary-strip {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-top: 12px; }}
    .mini {{ padding: 12px; border: 1px solid var(--line); border-radius: 8px; background: #fbfcfe; }}
    .mini .label {{ color: var(--muted); font-size: 12px; }}
    .mini .value {{ font-size: 19px; font-weight: 800; margin-top: 5px; }}
    .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }}
    .small-table {{ min-width: 0; }}
    .small-table table {{ min-width: 620px; }}
    .foot {{ color: var(--muted); font-size: 12px; margin-top: 10px; }}
    @media (max-width: 1200px) {{
      .kpis {{ grid-template-columns: repeat(3, 1fr); }}
      .grid, .two-col, .charts {{ grid-template-columns: 1fr; }}
    }}
    @media (max-width: 760px) {{
      header, main {{ padding-left: 14px; padding-right: 14px; }}
      .header-inner {{ align-items: flex-start; flex-direction: column; }}
      .kpis {{ grid-template-columns: 1fr 1fr; }}
      .rank-row {{ grid-template-columns: 1fr 80px; }}
      .rank-track {{ grid-column: 1 / -1; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="header-inner">
      <div class="brand">
        <div class="logo">PD</div>
        <div>
          <h1>상품DATA 운영 대시보드</h1>
          <div class="subtitle">자사몰 2026년 판매 데이터 기준 · 상품DATA 표준명으로 통합</div>
        </div>
      </div>
      <div class="status-pill">마지막 생성: {esc(updated_at)}</div>
    </div>
  </header>
  <main>
    <section class="kpis">
      <div class="kpi"><div class="label">자사몰 판매수량</div><div class="value">{money(total_qty)}</div><div class="note">세트 분해 반영</div></div>
      <div class="kpi"><div class="label">자사몰 실판매금액</div><div class="value">{money(total_payment)}</div><div class="note">결제금액 배분액</div></div>
      <div class="kpi"><div class="label">평균 판매단가</div><div class="value">{money(avg_unit)}</div><div class="note">실판매금액 / 수량</div></div>
      <div class="kpi"><div class="label">평균 할인율</div><div class="value">{pct(avg_discount)}</div><div class="note">정상가 기준</div></div>
      <div class="kpi"><div class="label">매칭률</div><div class="value">{pct(match_rate)}</div><div class="note">{matched}/{len(data)} 상품옵션</div></div>
      <div class="kpi"><div class="label">미매칭</div><div class="value">{unmatched}</div><div class="note">검토 필요 항목</div></div>
    </section>

    <nav class="tabs">
      <button class="tab-btn active" data-tab="overview">요약</button>
      <button class="tab-btn" data-tab="calendar">일자별 매출</button>
      <button class="tab-btn" data-tab="retailer">유통사별 매출</button>
      <button class="tab-btn" data-tab="detail">상품별 매출</button>
    </nav>

    <div id="overview" class="tab-panel active">
    <section class="grid">
      <div class="panel">
        <h2>판매 추이 요약</h2>
        <div class="charts">
          <div class="chart-box"><canvas id="paymentChart"></canvas></div>
          <div class="chart-box"><canvas id="qtyChart"></canvas></div>
        </div>
      </div>
      <div class="panel">
        <h2>결제금액 상위 상품</h2>
        {top_payment_rows}
      </div>
    </section>

    <section class="two-col">
      <div class="panel">
        <h2>업데이트 상태</h2>
        <div class="summary-strip">
          <div class="mini"><div class="label">상품DATA 품번 수</div><div class="value">{money(wb_summary['main_count'])}</div></div>
          <div class="mini"><div class="label">자사몰 집계행</div><div class="value">{money(len(data))}</div></div>
          <div class="mini"><div class="label">총 재고수량</div><div class="value">{money(wb_summary['original_inventory_qty'])}</div></div>
        </div>
        <div class="summary-strip">
          <div class="mini"><div class="label">단품/분해 구분</div><div class="value">{esc(', '.join(f'{k} {v["qty"]:,}' for k, v in by_type.items()))}</div></div>
          <div class="mini"><div class="label">정상가 기준 금액</div><div class="value">{money(total_gross)}</div></div>
          <div class="mini"><div class="label">자사몰 주문수</div><div class="value">{money(wb_summary['self_summary'].get('주문 수') or 0)}</div></div>
        </div>
        <div class="foot">신규 유통사 파일이 추가되면 같은 구조로 유통사별 금액을 별도 보존하고 통합 테이블에 추가할 수 있습니다.</div>
      </div>
      <div class="panel small-table">
        <h2>미매칭 검토</h2>
        <div class="table-wrap" style="max-height:270px">
          <table>
            <thead><tr><th>상품번호</th><th>표준상품명</th><th class="num">수량</th><th class="num">금액</th></tr></thead>
            <tbody>{unmatched_preview or '<tr><td colspan="4">미매칭 항목 없음</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </section>
    </div>

    <div id="calendar" class="tab-panel">
      <section class="panel">
        <h2>일자별 매출</h2>
        <div class="toolbar">
          <input id="dateStart" type="date" />
          <input id="dateEnd" type="date" />
          <button class="tab-btn active" id="clearDate" type="button">전체</button>
        </div>
        <div class="summary-strip">
          <div class="mini"><div class="label">선택 기간 매출</div><div class="value" id="dayPayment">0</div></div>
          <div class="mini"><div class="label">선택 기간 주문수</div><div class="value" id="dayOrders">0</div></div>
          <div class="mini"><div class="label">선택 기간 판매수량</div><div class="value" id="dayQty">0</div></div>
        </div>
        <div class="chart-box" style="margin-top:14px"><canvas id="dailyChart"></canvas></div>
        <div class="table-wrap" style="max-height:360px; margin-top:14px">
          <table>
            <thead><tr><th>일자</th><th class="num">매출</th><th class="num">정상가</th><th class="num">주문수</th><th class="num">판매수량</th><th class="num">객단가</th></tr></thead>
            <tbody id="dailyRows"></tbody>
          </table>
        </div>
      </section>
    </div>

    <div id="retailer" class="tab-panel">
      <section class="panel">
        <h2>유통사별 매출</h2>
        <div class="toolbar">
          <input id="retailerStart" type="date" />
          <input id="retailerEnd" type="date" />
          <select id="retailerFilter"><option value="">전체 유통사</option>{retailer_options}</select>
          <button class="tab-btn active" id="clearRetailer" type="button">전체</button>
        </div>
        <div class="chart-box"><canvas id="retailerChart"></canvas></div>
        <div class="table-wrap" style="max-height:360px; margin-top:14px">
          <table>
            <thead><tr><th>유통사</th><th class="num">매출</th><th class="num">정상가</th><th class="num">주문수</th><th class="num">판매수량</th><th class="num">객단가</th></tr></thead>
            <tbody id="retailerRows"></tbody>
          </table>
        </div>
        <div class="foot">현재 입력된 유통사는 자사몰입니다. 다른 유통사 파일이 들어오면 동일한 구조로 행이 추가됩니다.</div>
      </section>
    </div>

    <div id="detail" class="tab-panel">
    <section class="panel">
      <h2>상품별 매출</h2>
      <div class="toolbar">
        <input id="q" placeholder="상품명, 상품번호, 색상 검색" />
        <input id="productStart" type="date" />
        <input id="productEnd" type="date" />
        <select id="productRetailer"><option value="">전체 유통사</option>{retailer_options}</select>
        <select id="status"><option value="">전체 매칭상태</option><option>매칭됨</option><option>미매칭</option></select>
        <select id="sourceType"><option value="">전체 구분</option><option>단품</option><option>세트분해</option><option>팩분해</option></select>
        <select id="sortBy"><option value="payment">금액순</option><option value="qty">수량순</option><option value="name">상품명순</option></select>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>유통사</th><th>상품번호</th><th>표준상품명</th><th>자사몰상품명</th><th>구분</th><th>매칭</th>
              <th class="num">판매수량</th><th class="num">정상가금액</th><th class="num">실판매금액</th><th class="num">평균단가</th><th class="num">할인율</th><th class="num">판매율</th><th class="num">재고 소진율</th>
            </tr>
          </thead>
          <tbody id="detailRows"></tbody>
        </table>
      </div>
    </section>
    </div>
  </main>
  <script>
    const rawRows = {table_data};
    const productRows = {product_data};
    const unmatchedRows = {unmatched_data};
    const charts = {chart_data};
    const dailySales = {daily_data};
    const retailerSales = {retailer_data};

    const fmt = new Intl.NumberFormat('ko-KR');
    const detailRows = document.getElementById('detailRows');
    const q = document.getElementById('q');
    const status = document.getElementById('status');
    const sourceType = document.getElementById('sourceType');
    const sortBy = document.getElementById('sortBy');
    const dailyRows = document.getElementById('dailyRows');
    const retailerRows = document.getElementById('retailerRows');
    const dateStart = document.getElementById('dateStart');
    const dateEnd = document.getElementById('dateEnd');
    const clearDate = document.getElementById('clearDate');
    const retailerStart = document.getElementById('retailerStart');
    const retailerEnd = document.getElementById('retailerEnd');
    const retailerFilter = document.getElementById('retailerFilter');
    const clearRetailer = document.getElementById('clearRetailer');
    const productStart = document.getElementById('productStart');
    const productEnd = document.getElementById('productEnd');
    const productRetailer = document.getElementById('productRetailer');
    let dailyChart;
    let retailerChart;

    function escapeHtml(value) {{
      return String(value ?? '').replace(/[&<>"']/g, m => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[m]));
    }}

    function rowDiscount(row) {{
      const gross = Number(row.gross || 0);
      const payment = Number(row.payment || 0);
      return gross ? 1 - payment / gross : 0;
    }}

    function inDateRange(date, start, end) {{
      return (!start || date >= start) && (!end || date <= end);
    }}

    function summarizeDays(days, start, end) {{
      return (days || []).filter(day => inDateRange(day.date, start, end)).reduce((sum, day) => {{
        sum.qty += Number(day.qty || 0);
        sum.gross += Number(day.gross || 0);
        sum.payment += Number(day.payment || 0);
        sum.orders += Number(day.orders || 0);
        return sum;
      }}, {{ qty: 0, gross: 0, payment: 0, orders: 0 }});
    }}

    function periodRow(row, start, end) {{
      if (!start && !end) return {{
        qty: Number(row.qty || 0),
        gross: Number(row.gross || 0),
        payment: Number(row.payment || 0),
        orders: Number(row.orders || 0),
      }};
      return summarizeDays(row.daily || [], start, end);
    }}

    function rateText(numerator, denominator) {{
      const bottom = Number(denominator || 0);
      return bottom ? `${{((Number(numerator || 0) / bottom) * 100).toFixed(1)}}%` : '-';
    }}

    function renderTable() {{
      const query = q.value.trim().toLowerCase();
      const st = status.value;
      const typ = sourceType.value;
      const sorter = sortBy.value;
      const start = productStart.value;
      const end = productEnd.value;
      const mall = productRetailer.value;
      let rows = rawRows.map(row => ({{ ...row, period: periodRow(row, start, end) }})).filter(row => {{
        const text = `${{row.mall_no}} ${{row.standard_name || row.name}} ${{row.name}} ${{row.color}} ${{row.size}}`.toLowerCase();
        return (!query || text.includes(query))
          && (!st || row.match_status === st)
          && (!typ || row.source_type === typ)
          && (!mall || (row.retailer || '자사몰') === mall)
          && ((!start && !end) || Number(row.period.qty || 0) > 0);
      }});
      rows.sort((a, b) => {{
        if (sorter === 'qty') return Number(b.period.qty || 0) - Number(a.period.qty || 0);
        if (sorter === 'name') return String(a.standard_name || a.name).localeCompare(String(b.standard_name || b.name), 'ko');
        return Number(b.period.payment || 0) - Number(a.period.payment || 0);
      }});
      detailRows.innerHTML = rows.slice(0, 500).map(row => {{
        const cls = row.match_status === '매칭됨' ? 'matched' : 'unmatched';
        const qty = Number(row.period.qty || 0);
        const gross = Number(row.period.gross || 0);
        const payment = Number(row.period.payment || 0);
        const avgUnit = qty ? Math.round(payment / qty) : 0;
        const periodDiscount = gross ? 1 - payment / gross : 0;
        const receivedQty = Number(row.received_qty || 0);
        const stockQty = Number(row.stock_qty || 0);
        return `<tr>
          <td>${{escapeHtml(row.retailer || '자사몰')}}</td>
          <td>${{escapeHtml(row.mall_no)}}</td>
          <td>${{escapeHtml(row.standard_name || row.name)}}</td>
          <td>${{escapeHtml(row.name)}}</td>
          <td><span class="badge source">${{escapeHtml(row.source_type || '단품')}}</span></td>
          <td><span class="badge ${{cls}}">${{escapeHtml(row.match_status)}}</span></td>
          <td class="num">${{fmt.format(qty)}}</td>
          <td class="num">${{fmt.format(gross)}}</td>
          <td class="num">${{fmt.format(payment)}}</td>
          <td class="num">${{fmt.format(avgUnit)}}</td>
          <td class="num">${{(periodDiscount * 100).toFixed(1)}}%</td>
          <td class="num">${{rateText(qty, receivedQty)}}</td>
          <td class="num">${{rateText(qty, qty + stockQty)}}</td>
        </tr>`;
      }}).join('');
    }}

    [q, status, sourceType, sortBy, productStart, productEnd, productRetailer].forEach(el => el.addEventListener('input', renderTable));
    renderTable();

    function renderDaily() {{
      const start = dateStart.value;
      const end = dateEnd.value;
      const rows = dailySales.filter(row => inDateRange(row.date, start, end));
      const payment = rows.reduce((sum, row) => sum + Number(row.payment || 0), 0);
      const orders = rows.reduce((sum, row) => sum + Number(row.orders || 0), 0);
      const qty = rows.reduce((sum, row) => sum + Number(row.qty || 0), 0);
      document.getElementById('dayPayment').textContent = fmt.format(payment);
      document.getElementById('dayOrders').textContent = fmt.format(orders);
      document.getElementById('dayQty').textContent = fmt.format(qty);
      dailyRows.innerHTML = rows.slice().reverse().map(row => `<tr>
        <td>${{escapeHtml(row.date)}}</td>
        <td class="num">${{fmt.format(row.payment || 0)}}</td>
        <td class="num">${{fmt.format(row.gross || 0)}}</td>
        <td class="num">${{fmt.format(row.orders || 0)}}</td>
        <td class="num">${{fmt.format(row.qty || 0)}}</td>
        <td class="num">${{fmt.format(row.avgOrder || 0)}}</td>
      </tr>`).join('');
      if (dailyChart) {{
        dailyChart.data.labels = rows.map(row => row.date);
        dailyChart.data.datasets[0].data = rows.map(row => row.payment);
        dailyChart.update();
      }}
    }}

    function retailerPeriodRows() {{
      const start = retailerStart.value;
      const end = retailerEnd.value;
      const selected = retailerFilter.value;
      const grouped = new Map();
      rawRows.forEach(row => {{
        const mall = row.retailer || '자사몰';
        if (selected && mall !== selected) return;
        const period = periodRow(row, start, end);
        if (!period.qty && (start || end)) return;
        const item = grouped.get(mall) || {{ retailer: mall, payment: 0, gross: 0, orders: 0, qty: 0 }};
        item.payment += Number(period.payment || 0);
        item.gross += Number(period.gross || 0);
        item.orders += Number(period.orders || 0);
        item.qty += Number(period.qty || 0);
        grouped.set(mall, item);
      }});
      return Array.from(grouped.values()).map(row => ({{
        ...row,
        avgOrder: row.orders ? Math.round(row.payment / row.orders) : 0,
      }})).sort((a, b) => b.payment - a.payment);
    }}

    function renderRetailer() {{
      const rows = retailerPeriodRows();
      retailerRows.innerHTML = rows.map(row => `<tr>
        <td>${{escapeHtml(row.retailer)}}</td>
        <td class="num">${{fmt.format(row.payment || 0)}}</td>
        <td class="num">${{fmt.format(row.gross || 0)}}</td>
        <td class="num">${{fmt.format(row.orders || 0)}}</td>
        <td class="num">${{fmt.format(row.qty || 0)}}</td>
        <td class="num">${{fmt.format(row.avgOrder || 0)}}</td>
      </tr>`).join('');
      if (retailerChart) {{
        retailerChart.data.labels = rows.map(row => row.retailer);
        retailerChart.data.datasets[0].data = rows.map(row => row.payment);
        retailerChart.update();
      }}
    }}

    [dateStart, dateEnd].forEach(el => el.addEventListener('input', renderDaily));
    clearDate.addEventListener('click', () => {{ dateStart.value = ''; dateEnd.value = ''; renderDaily(); }});
    [retailerStart, retailerEnd, retailerFilter].forEach(el => el.addEventListener('input', renderRetailer));
    clearRetailer.addEventListener('click', () => {{ retailerStart.value = ''; retailerEnd.value = ''; retailerFilter.value = ''; renderRetailer(); }});
    renderDaily();
    renderRetailer();

    document.querySelectorAll('.tab-btn[data-tab]').forEach(btn => {{
      btn.addEventListener('click', () => {{
        document.querySelectorAll('.tab-btn[data-tab]').forEach(item => item.classList.remove('active'));
        document.querySelectorAll('.tab-panel').forEach(item => item.classList.remove('active'));
        btn.classList.add('active');
        document.getElementById(btn.dataset.tab).classList.add('active');
      }});
    }});

    Chart.defaults.font.family = "Arial, 'Malgun Gothic', sans-serif";
    Chart.defaults.color = '#475467';
    new Chart(document.getElementById('paymentChart'), {{
      type: 'line',
      data: {{
        labels: charts.paymentLabels,
        datasets: [{{ label: '일자별 매출', data: charts.paymentValues, borderColor: '#1f5f99', backgroundColor: 'rgba(31,95,153,.12)', fill: true, tension: .25 }}]
      }},
      options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }}, scales: {{ x: {{ ticks: {{ maxRotation: 45, minRotation: 0, callback: function(value, index) {{ return index % 3 === 0 ? this.getLabelForValue(value) : ''; }} }} }}, y: {{ ticks: {{ callback: value => fmt.format(value) }} }} }} }}
    }});
    new Chart(document.getElementById('qtyChart'), {{
      type: 'bar',
      data: {{
        labels: charts.qtyLabels,
        datasets: [{{ label: `주간 판매수량 (${{charts.qtyWeekLabel}})`, data: charts.qtyValues, backgroundColor: '#0f766e', borderRadius: 4 }}]
      }},
      options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: true }} }}, scales: {{ x: {{ ticks: {{ maxRotation: 45, minRotation: 0 }} }}, y: {{ ticks: {{ callback: value => fmt.format(value) }} }} }} }}
    }});
    dailyChart = new Chart(document.getElementById('dailyChart'), {{
      type: 'line',
      data: {{
        labels: dailySales.map(row => row.date),
        datasets: [{{ label: '일자별 매출', data: dailySales.map(row => row.payment), borderColor: '#1f5f99', backgroundColor: 'rgba(31,95,153,.12)', fill: true, tension: .25 }}]
      }},
      options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ ticks: {{ callback: value => fmt.format(value) }} }} }} }}
    }});
    retailerChart = new Chart(document.getElementById('retailerChart'), {{
      type: 'bar',
      data: {{
        labels: retailerSales.map(row => row.retailer),
        datasets: [{{ label: '유통사별 매출', data: retailerSales.map(row => row.payment), backgroundColor: '#0f766e', borderRadius: 4 }}]
      }},
      options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ ticks: {{ callback: value => fmt.format(value) }} }} }} }}
    }});
  </script>
</body>
</html>"""
    HTML_OUT.write_text(html, encoding="utf-8")
    print(HTML_OUT)


if __name__ == "__main__":
    make_dashboard()
