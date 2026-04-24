import json
import re
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import openpyxl


OUT_FILE = Path("historical_daily.json")

FILE_2025 = Path(r"F:/★ 2025 일일판매정리(12월).xlsx")
FILE_2024 = Path(r"\\192.168.0.4\woojoo\플래니트\1.매출정리, 프로모션\2024 일일판매정리\★2024_일일판매정리_24.12.31.xlsx")
FILE_2023 = Path(r"\\192.168.0.4\woojoo\플래니트\1.매출정리, 프로모션\2023 일일판매정리\★2023_일일판매정리.xlsx")


def normalize_retailer(value):
    if value is None:
        return None
    text = str(value).replace("\n", "").replace(" ", "").strip().upper()
    if not text:
        return None
    if text.startswith("주간총합") or "총합" in text or "촬영관련" in text:
        return None
    if "자사몰" in text:
        return "자사몰"
    if "무신사" in text:
        return "무신사"
    if "4XR" in text:
        return "4XR"
    if "29CM" in text:
        return "29cm"
    if "애슬러" in text and "롯데온" in text:
        return None
    if "애슬러" in text:
        return "애슬러"
    if "롯데온" in text:
        return "롯데온"
    if "네이버" in text:
        return "네이버"
    return None


def as_int(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    return None


def extract_day_num(value):
    if isinstance(value, (int, float)):
        day = int(value)
        return day if 1 <= day <= 31 else None
    if isinstance(value, str):
        m = re.search(r"(\d{1,2})", value)
        if m:
            day = int(m.group(1))
            return day if 1 <= day <= 31 else None
    return None


def has_numeric(values):
    return any(isinstance(v, (int, float)) for v in values if v is not None)


def add_amount(bucket, year, dt, retailer, amount):
    if amount is None:
        return
    if dt.year != year:
        return
    day = bucket[year].setdefault(
        dt.isoformat(),
        {"date": dt.isoformat(), "payment": 0, "retailers": defaultdict(int)},
    )
    day["payment"] += int(amount)
    if retailer:
        day["retailers"][retailer] += int(amount)


def finalize(bucket):
    out = {}
    for year, days in bucket.items():
        out[str(year)] = []
        for key in sorted(days):
            row = days[key]
            out[str(year)].append(
                {
                    "date": row["date"],
                    "payment": row["payment"],
                    "retailers": dict(sorted(row["retailers"].items())),
                }
            )
    return out


def parse_2024_style_sheet(path, sheet_name, year, start_date):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    bucket = defaultdict(dict)
    next_expected = start_date
    week_dates = []

    for idx, row in enumerate(rows):
        day_nums = [extract_day_num(v) for v in row[1:8]]
        valid_days = [d for d in day_nums if d and 1 <= d <= 31]
        if len(valid_days) >= 3 and str(row[8] or "").strip() == "주간 실매출":
            week_dates = []
            for day_num in day_nums:
                if not day_num:
                    week_dates.append(None)
                    continue
                guard = 0
                while next_expected.day != day_num and guard < 40:
                    next_expected += timedelta(days=1)
                    guard += 1
                week_dates.append(next_expected)
                next_expected += timedelta(days=1)
            continue

        retailer = normalize_retailer(row[0] if row else None)
        if not retailer or not week_dates:
            continue

        amount_row = row[1:8]
        if not has_numeric(amount_row):
            for look_ahead in range(1, 5):
                if idx + look_ahead >= len(rows):
                    break
                next_row = rows[idx + look_ahead]
                next_label = normalize_retailer(next_row[0] if next_row else None)
                next_header_days = [extract_day_num(v) for v in next_row[1:8]]
                if next_label or (len([d for d in next_header_days if d]) >= 3 and str(next_row[8] or "").strip() == "주간 실매출"):
                    break
                if not next_row[0] and has_numeric(next_row[1:8]):
                    amount_row = next_row[1:8]
                    break

        for pos, cell in enumerate(amount_row):
            dt = week_dates[pos] if pos < len(week_dates) else None
            amount = as_int(cell)
            if dt and amount is not None:
                add_amount(bucket, year, dt, retailer, amount)

    return finalize(bucket)


def parse_2023_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["일별매출2023"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    bucket = defaultdict(dict)
    week_dates = []
    date_cols = []
    rx = re.compile(r"(\d{1,2})/(\d{1,2})")

    for row in rows:
        current_cols = []
        current_dates = []
        for i, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            m = rx.search(cell)
            if not m:
                continue
            month = int(m.group(1))
            day = int(m.group(2))
            current_cols.append(i)
            current_dates.append(date(2023, month, day))

        if len(current_dates) >= 3:
            date_cols = current_cols
            week_dates = current_dates
            continue

        retailer = normalize_retailer(row[0] if row else None)
        if not retailer or not week_dates or not date_cols:
            continue

        for pos, col_idx in enumerate(date_cols):
            sales_idx = col_idx + 2
            if sales_idx >= len(row):
                continue
            amount = as_int(row[sales_idx])
            if amount is not None:
                add_amount(bucket, 2023, week_dates[pos], retailer, amount)

    return finalize(bucket)


def merge_year_data(target, source):
    for year, rows in source.items():
        year_bucket = {r["date"]: r for r in target.setdefault(year, [])}
        for row in rows:
            if row["date"] not in year_bucket:
                year_bucket[row["date"]] = {"date": row["date"], "payment": 0, "retailers": {}}
            year_bucket[row["date"]]["payment"] += row["payment"]
            for retailer, amount in row["retailers"].items():
                year_bucket[row["date"]]["retailers"][retailer] = year_bucket[row["date"]]["retailers"].get(retailer, 0) + amount
        target[year] = [year_bucket[k] for k in sorted(year_bucket)]


def main():
    data = {}
    merge_year_data(data, parse_2023_sheet(FILE_2023))
    merge_year_data(data, parse_2024_style_sheet(FILE_2024, "일별매출정리", 2024, date(2024, 1, 1)))
    merge_year_data(data, parse_2024_style_sheet(FILE_2025, "일별매출정리", 2025, date(2025, 12, 1)))

    OUT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    for year, rows in sorted(data.items()):
        total = sum(r["payment"] for r in rows)
        print(year, len(rows), total, rows[0]["date"] if rows else "-", rows[-1]["date"] if rows else "-")


if __name__ == "__main__":
    main()
