import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_FILE = ROOT / "data.json"
UNMATCHED_FILE = ROOT / "stock_unmatched.json"
DESKTOP = Path.home() / "Desktop"


def normalize(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def normalize_name(value) -> str:
    text = normalize(value)
    text = re.sub(r"^\[[^\]]+\]", "", text)
    return re.sub(r"[-_()\[\]/,]", "", text)


SIZE_PATTERN = re.compile(r"(?:^|[^A-Z0-9])(4XL|3XL|2XL|XXXL|XXL|XL|XS|FREE|F|S|M|L)(?=$|[^A-Z0-9])", re.I)
SIZE_SUFFIX_PATTERN = re.compile(r"(4XL|3XL|2XL|XXXL|XXL|XL|XS|FREE|F|S|M|L)$", re.I)
COLOR_ALIASES = {
    "BLACK": "BLACK",
    "BLK": "BLACK",
    "블랙": "BLACK",
    "WHITE": "WHITE",
    "WHT": "WHITE",
    "화이트": "WHITE",
    "NAVY": "NAVY",
    "NVY": "NAVY",
    "네이비": "NAVY",
    "CHARCOAL": "CHARCOAL",
    "차콜": "CHARCOAL",
    "GREY": "GREY",
    "GRAY": "GREY",
    "GRY": "GREY",
    "LGR": "GREY",
    "MGREY": "GREY",
    "LIGHTGREY": "GREY",
    "LIGHTGRAY": "GREY",
    "라이트그레이": "GREY",
    "MELANGEGREY": "GREY",
    "멜란지그레이": "GREY",
    "KHAKI": "KHAKI",
    "LKHAKI": "KHAKI",
    "카키": "KHAKI",
    "BEIGE": "BEIGE",
    "베이지": "BEIGE",
    "BROWN": "BROWN",
    "브라운": "BROWN",
}
COLOR_SUFFIXES = sorted(COLOR_ALIASES.keys(), key=len, reverse=True)


def normalize_size(value) -> str:
    text = str(value or "").upper().replace("(", " ").replace(")", " ")
    matches = SIZE_PATTERN.findall(text)
    if not matches:
        return ""
    size = matches[-1].upper()
    return "FREE" if size == "F" else size


def product_key(value) -> str:
    text = normalize_name(value)
    if not text:
        return ""
    text = SIZE_SUFFIX_PATTERN.sub("", text)
    for color in COLOR_SUFFIXES:
        text = re.sub(re.escape(color) + r"$", "", text, flags=re.I)
    return text


def color_key(*values) -> str:
    text = "".join(normalize_name(value) for value in values if value)
    colors = {canonical for alias, canonical in COLOR_ALIASES.items() if alias in text}
    return "|".join(sorted(colors))


def latest_stock_file() -> Path:
    patterns = [
        "재고조회(기본)_*.xlsx",
        "재고조회_*.xlsx",
        "*재고조회*.xlsx",
    ]
    files: list[Path] = []
    for pattern in patterns:
        files.extend([p for p in DESKTOP.glob(pattern) if not p.name.startswith("~$")])
    if not files:
        raise FileNotFoundError("재고조회(기본) 엑셀 파일을 바탕화면에서 찾지 못했습니다.")
    return max(files, key=lambda p: p.stat().st_mtime)


def load_stock_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [
        str(v).strip() if v is not None else ""
        for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    idx = {h: i for i, h in enumerate(headers)}

    if "출고가능" in idx:
        stock_column = "출고가능"
    elif "현재고" in idx:
        stock_column = "현재고"
    else:
        stock_column = "총재고"

    required = ["출고상품명", "바코드", stock_column]
    missing = [h for h in required if h not in idx]
    if missing:
        raise ValueError(f"필수 컬럼 누락: {missing}")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        barcode = str(row[idx["바코드"]] or "").strip()
        outbound_name = str(row[idx["출고상품명"]] or "").strip()
        stock_raw = row[idx[stock_column]]
        try:
            stock_qty = int(float(str(stock_raw or 0).replace(",", "")))
        except ValueError:
            stock_qty = 0
        if not barcode and not outbound_name:
            continue
        rows.append(
            {
                "barcode": barcode,
                "outbound_name": outbound_name,
                "stock_qty": stock_qty,
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("stock_file", nargs="?", help="재고조회(기본) 엑셀 파일 경로")
    args = parser.parse_args()

    stock_file = Path(args.stock_file) if args.stock_file else latest_stock_file()
    data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    stock_rows = load_stock_rows(stock_file)

    merged_by_name: dict[str, dict] = {}
    for r in stock_rows:
        key = normalize(r.get("outbound_name"))
        if not key:
            continue
        if key not in merged_by_name:
            merged_by_name[key] = dict(r)
        else:
            merged_by_name[key]["stock_qty"] = int(merged_by_name[key].get("stock_qty") or 0) + int(
                r.get("stock_qty") or 0
            )
            if not merged_by_name[key].get("barcode") and r.get("barcode"):
                merged_by_name[key]["barcode"] = r.get("barcode")
    stock_rows = list(merged_by_name.values())

    by_barcode = {normalize(r["barcode"]): r for r in stock_rows if r["barcode"]}
    by_name = {normalize(r["outbound_name"]): r for r in stock_rows if r["outbound_name"]}
    by_product_size: dict[str, dict] = {}
    product_colors: dict[str, set[str]] = {}
    for r in stock_rows:
        pkey = product_key(r.get("outbound_name"))
        if not pkey:
            continue
        row_color = color_key(r.get("outbound_name"), r.get("barcode"))
        if row_color:
            product_colors.setdefault(pkey, set()).add(row_color)
        row_size = normalize_size(r.get("outbound_name"))
        if not row_size:
            continue
        product_size_key = f"{pkey}|{row_size}"
        if product_size_key not in by_product_size:
            by_product_size[product_size_key] = {
                "barcode": r.get("barcode") or "",
                "outbound_name": r.get("outbound_name") or "",
                "stock_qty": 0,
                "used_keys": set(),
            }
        aggregate = by_product_size[product_size_key]
        aggregate["stock_qty"] += int(r.get("stock_qty") or 0)
        if not aggregate.get("barcode") and r.get("barcode"):
            aggregate["barcode"] = r.get("barcode")
        for used_key in (normalize(r.get("barcode")), normalize(r.get("outbound_name"))):
            if used_key:
                aggregate["used_keys"].add(used_key)

    stock_name_pairs = [
        (stock_name, r)
        for r in stock_rows
        if r["outbound_name"]
        for stock_name in [normalize_name(r["outbound_name"])]
        if len(stock_name) >= 6
    ]

    matched = 0
    unmatched_data: list[dict] = []
    used_stock_keys: set[str] = set()

    for item in data:
        candidates = [
            normalize(item.get("match_sku")),
            normalize(item.get("standard_name")),
            normalize(item.get("name")),
        ]
        stock_row = None
        match_key = None
        item_color = str(item.get("color") or "").strip()
        item_product_key = product_key(item.get("name")) or product_key(item.get("standard_name"))
        item_size = normalize_size(item.get("size")) or normalize_size(item.get("standard_name"))
        product_size_key = f"{item_product_key}|{item_size}" if item_product_key and item_size else ""
        if (
            product_size_key
            and (not item_color or len(product_colors.get(item_product_key, set())) <= 1)
            and product_size_key in by_product_size
        ):
            stock_row = by_product_size[product_size_key]
            match_key = product_size_key
        for key in candidates:
            if stock_row is not None:
                break
            if key and key in by_barcode:
                stock_row = by_barcode[key]
                match_key = key
                break
        if stock_row is None:
            for key in candidates:
                if key and key in by_name:
                    stock_row = by_name[key]
                    match_key = key
                    break
        if stock_row is None:
            name_candidates = [
                normalize_name(item.get("standard_name")),
                normalize_name(item.get("name")),
            ]
            name_candidates = [v for v in name_candidates if len(v) >= 6]
            matched_by_name: list[dict] = []
            for candidate in name_candidates:
                matched_by_name = [
                    r
                    for stock_name, r in stock_name_pairs
                    if candidate in stock_name or stock_name in candidate
                ]
                if matched_by_name:
                    break
            if matched_by_name:
                stock_row = {
                    "barcode": "",
                    "outbound_name": matched_by_name[0]["outbound_name"],
                    "stock_qty": matched_by_name[0]["stock_qty"],
                }
                match_key = normalize_name(stock_row["outbound_name"])

        if stock_row:
            item["stock_qty"] = stock_row["stock_qty"]
            item["stock_barcode"] = stock_row["barcode"]
            item["stock_name"] = stock_row["outbound_name"]
            matched += 1
            used_keys = stock_row.get("used_keys") if isinstance(stock_row, dict) else None
            if used_keys:
                for used_key in used_keys:
                    used_stock_keys.add(used_key)
            elif match_key:
                used_stock_keys.add(match_key)
        else:
            item["stock_qty"] = 0
            item["stock_barcode"] = ""
            item["stock_name"] = ""
            unmatched_data.append(
                {
                    "retailer": item.get("retailer"),
                    "match_sku": item.get("match_sku"),
                    "standard_name": item.get("standard_name"),
                    "name": item.get("name"),
                    "color": item.get("color"),
                    "size": item.get("size"),
                }
            )

    unmatched_stock = [
        r
        for r in stock_rows
        if normalize(r["barcode"]) not in used_stock_keys
        and normalize(r["outbound_name"]) not in used_stock_keys
    ]

    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    UNMATCHED_FILE.write_text(
        json.dumps(
            {
                "stock_file": str(stock_file),
                "data_rows": len(data),
                "stock_rows": len(stock_rows),
                "matched_data_rows": matched,
                "unmatched_data_rows": unmatched_data,
                "unmatched_stock_rows": unmatched_stock,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "stock_file": str(stock_file),
                "data_rows": len(data),
                "stock_rows": len(stock_rows),
                "matched_data_rows": matched,
                "unmatched_data_rows": len(unmatched_data),
                "unmatched_stock_rows": len(unmatched_stock),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()

