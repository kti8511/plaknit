import argparse
import collections
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl


DATA_FILE = Path("data.json")
MANUAL_FILE = Path("manual_sales_updates.json")
RETAILER = "무신사"


SIZE_ALIASES = {
    "XS": "XS",
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "2XL",
    "XXXL": "3XL",
    "XXXXL": "4XL",
    "2XL": "2XL",
    "3XL": "3XL",
    "4XL": "4XL",
    "FREE": "FREE",
}

COLOR_ALIASES = {
    "BLACK": "블랙",
    "WHITE": "화이트",
    "CHARCOAL": "차콜",
    "DEEP CHARCOAL": "차콜",
    "GRAY": "그레이",
    "GREY": "그레이",
    "LIGHT GRAY": "라이트그레이",
    "LIGHT GREY": "라이트그레이",
    "LIGHTGRAY": "라이트그레이",
    "LIGHTGREY": "라이트그레이",
    "MELANGE GRAY": "멜란지그레이",
    "MELANGE GREY": "멜란지그레이",
    "NAVY": "네이비",
    "BLUE": "블루",
    "ORANGE": "오렌지",
    "BEIGE": "베이지",
    "BROWN": "브라운",
    "KHAKI": "카키",
    "KHAKI GRAY": "카키그레이",
    "KHAKI GREY": "카키그레이",
    "VINTAGE KHAKI": "카키",
}


NAME_ALIASES = {
    "KINTERRA 퀵드라이 스포츠 나시": "KINTERRA 퀵 드라이 스포츠 나시",
    "KINTERRA 퀵드라이 스포츠 오버핏 반팔티": "KINTERRA 퀵 드라이 오버핏 반팔티",
    "KINTERRA 퀵드라이 스포츠 머슬핏 반팔티": "KINTERRA 퀵 드라이 스포츠 머슬핏 반팔티",
    "TECH SWEAT 피그먼트 빈티지 조거팬츠 (베이직)": "TECH SWEAT 피그먼트 소프트 빈티지 조거팬츠(베이직)",
    "[헤비쮸리] TECH SWEAT 피그먼트 빈티지 조거팬츠 (베이직)": "TECH SWEAT 피그먼트 소프트 빈티지 조거팬츠(베이직)",
    "[헤비쮸리] TECH SWEAT 피그먼트 빈티지 맨투맨 (세미오버핏)": "TECH SWEAT 피그먼트 빈티지 맨투맨 (세미오버핏)",
    "[플래니트X박영감] 빈티지 세미크롭 스웻셔츠 멜란지그레이": "[플래니트X박영감] 빈티지 워싱 스웻셔츠",
    "ICE LITE 멀티쇼츠(이너내장)": "ICE LITE 멀티쇼츠",
}


NAME_COLOR_SUFFIXES = [
    (" 라이트그레이", "라이트그레이"),
    (" 멜란지그레이", "멜란지그레이"),
    (" 카키그레이", "카키그레이"),
    (" 웜그레이", "웜그레이"),
    (" 라이트블루", "라이트블루"),
    (" 다크브라운", "다크브라운"),
    (" 블랙", "블랙"),
    (" 화이트", "화이트"),
    (" 네이비", "네이비"),
    (" 그레이", "그레이"),
    (" 차콜", "차콜"),
    (" 브라운", "브라운"),
    (" 세이지", "세이지"),
    (" 카키", "카키"),
    (" 레드", "레드"),
]


def clean_text(value):
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def norm_size(value):
    text = clean_text(value).upper()
    return SIZE_ALIASES.get(text, text)


def norm_color(value):
    text = clean_text(value)
    upper = text.upper()
    upper = upper.replace("  ", " ")
    if "," in text:
        return ", ".join(norm_color(part) for part in text.split(","))
    return COLOR_ALIASES.get(upper, text.replace("라이트 그레이", "라이트그레이"))


def normalize_name(name):
    text = clean_text(name)
    text = re.sub(r"^\*", "", text).strip()
    text = text.replace("퀵 드라이", "퀵드라이")
    text = text.replace("세미 오버핏", "세미오버핏")
    text = text.replace("세비오버핏", "세미오버핏")
    text = text.replace("초냉감 트래블러 티셔츠", "채코제에디션 초냉감 트래블러티셔츠")
    text = text.replace("[채코제에디션]", "채코제에디션 ").strip()
    text = text.replace("채코제 에디션", "채코제에디션")
    text = text.replace("차코제에디션", "채코제에디션")
    text = text.replace("7인치", '7"')
    text = NAME_ALIASES.get(text, text)
    return clean_text(text)


def split_name_color(name, color):
    name = normalize_name(name)
    color = norm_color(color)
    if not color:
        for suffix, suffix_color in NAME_COLOR_SUFFIXES:
            if name.endswith(suffix):
                return name[: -len(suffix)].strip(), suffix_color
    if name in NAME_ALIASES:
        return NAME_ALIASES[name], color
    return name, color


def parse_option(option):
    option = clean_text(option)
    if not option or ":" not in option:
        return "", ""
    left, right = [part.strip() for part in option.split(":", 1)]
    left_size = norm_size(left)
    right_size = norm_size(right)
    if left_size in SIZE_ALIASES.values():
        return norm_color(right), left_size
    if right_size in SIZE_ALIASES.values():
        return norm_color(left), right_size
    return norm_color(left), norm_size(right)


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).replace(".", "-")[:10]


def as_int(value):
    if value in (None, ""):
        return 0
    return int(round(float(str(value).replace(",", ""))))


def standard_name(name, color, size):
    if color and size:
        return f"{name}_{color}-{size}"
    if color:
        return f"{name}_{color}"
    if size:
        return f"{name}-{size}"
    return name


def row_option_key(row):
    name = normalize_name(row.get("name"))
    color = norm_color(row.get("color"))
    size = norm_size(row.get("size"))
    if (not color or not size) and row.get("standard_name"):
        std = clean_text(row.get("standard_name"))
        if "_" in std:
            before, after = std.rsplit("_", 1)
            if "-" in after:
                c, s = after.rsplit("-", 1)
                name = normalize_name(before)
                color = color or norm_color(c)
                size = size or norm_size(s)
            elif "_" in after:
                c, s = after.rsplit("_", 1)
                name = normalize_name(before)
                color = color or norm_color(c)
                size = size or norm_size(s)
    return name, color, size


def remove_existing_april(rows):
    for row in rows:
        if row.get("retailer") != RETAILER:
            continue
        kept = []
        removed_qty = removed_gross = removed_payment = removed_orders = 0
        for daily in row.get("daily", []):
            if str(daily.get("date", "")).startswith("2026-04-"):
                removed_qty += as_int(daily.get("qty"))
                removed_gross += as_int(daily.get("gross"))
                removed_payment += as_int(daily.get("payment"))
                removed_orders += as_int(daily.get("orders"))
            else:
                kept.append(daily)
        if len(kept) != len(row.get("daily", [])):
            row["daily"] = kept
            row["qty"] = as_int(row.get("qty")) - removed_qty
            row["gross"] = as_int(row.get("gross")) - removed_gross
            row["payment"] = as_int(row.get("payment")) - removed_payment
            row["orders"] = as_int(row.get("orders")) - removed_orders
            row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def build_indexes(rows):
    exact = collections.defaultdict(list)
    by_name = collections.defaultdict(list)
    for index, row in enumerate(rows):
        if row.get("retailer") != RETAILER:
            continue
        key = row_option_key(row)
        exact[key].append(index)
        by_name[key[0]].append(index)
    return exact, by_name


def find_match(exact, by_name, name, color, size):
    for key in ((name, color, size), (name, "", size), (name, color, ""), (name, "", "")):
        if exact.get(key):
            return exact[key][0]
    if by_name.get(name):
        return by_name[name][0]
    return None


def append_daily(row, item):
    row.setdefault("daily", []).append(
        {
            "date": item["date"],
            "qty": item["qty"],
            "gross": item["gross"],
            "payment": item["payment"],
            "orders": 1,
        }
    )
    row["qty"] = as_int(row.get("qty")) + item["qty"]
    row["gross"] = as_int(row.get("gross")) + item["gross"]
    row["payment"] = as_int(row.get("payment")) + item["payment"]
    row["orders"] = as_int(row.get("orders")) + 1
    row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def merge_daily(row):
    merged = {}
    for daily in row.get("daily", []):
        day = daily.get("date")
        current = merged.setdefault(day, {"date": day, "qty": 0, "gross": 0, "payment": 0, "orders": 0})
        current["qty"] += as_int(daily.get("qty"))
        current["gross"] += as_int(daily.get("gross"))
        current["payment"] += as_int(daily.get("payment"))
        current["orders"] += as_int(daily.get("orders"))
    row["daily"] = [merged[day] for day in sorted(merged)]


def load_items(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    items = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        row = dict(zip(headers, values))
        day = parse_date(row.get("주문일시"))
        if not day.startswith("2026-04-"):
            continue
        color, size = parse_option(row.get("옵션"))
        name, color = split_name_color(row.get("상품명"), color)
        qty = as_int(row.get("수량"))
        payment = as_int(row.get("매출금액"))
        gross = as_int(row.get("정상가")) * qty
        items.append(
            {
                "date": day,
                "name": name,
                "color": color,
                "size": size,
                "qty": qty,
                "gross": gross,
                "payment": payment,
                "status": clean_text(row.get("주문상태")),
                "raw_name": clean_text(row.get("상품명")),
                "raw_option": clean_text(row.get("옵션")),
            }
        )
    return items


def update_manual_sales(items):
    manual = json.loads(MANUAL_FILE.read_text(encoding="utf-8"))
    daily = collections.defaultdict(lambda: {"payment": 0, "qty": 0, "orders": 0})
    for item in items:
        daily[item["date"]]["payment"] += item["payment"]
        daily[item["date"]]["qty"] += item["qty"]
        daily[item["date"]]["orders"] += 1

    by_date = {entry["date"]: entry for entry in manual}
    for day, totals in daily.items():
        entry = by_date.setdefault(day, {"date": day, "retailers": []})
        retailers = entry.setdefault("retailers", [])
        target = next((r for r in retailers if r.get("retailer") == RETAILER), None)
        if target is None:
            target = {"retailer": RETAILER, "items": []}
            retailers.append(target)
        target["payment"] = totals["payment"]
        target["qty"] = totals["qty"]
        target["orders"] = totals["orders"]
        target["items"] = []
    return [by_date[day] for day in sorted(by_date)]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    items = load_items(Path(args.xlsx))
    rows = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    remove_existing_april(rows)
    exact, by_name = build_indexes(rows)
    created = []

    for item in items:
        match_idx = find_match(exact, by_name, item["name"], item["color"], item["size"])
        if match_idx is None:
            new_row = {
                "retailer": RETAILER,
                "mall_no": "",
                "name": item["name"],
                "color": item["color"],
                "size": item["size"],
                "qty": 0,
                "gross": 0,
                "payment": 0,
                "avg_unit": 0,
                "orders": 0,
                "source_type": "상품",
                "daily": [],
                "match_status": "매칭완료",
                "match_sku": "",
                "standard_name": standard_name(item["name"], item["color"], item["size"]),
                "received_qty": 0,
                "stock_qty": 0,
            }
            rows.append(new_row)
            match_idx = len(rows) - 1
            exact[(item["name"], item["color"], item["size"])].append(match_idx)
            by_name[item["name"]].append(match_idx)
            created.append(new_row["standard_name"])

        row = rows[match_idx]
        row["name"] = item["name"]
        row["color"] = item["color"]
        row["size"] = item["size"]
        row["standard_name"] = standard_name(item["name"], item["color"], item["size"])
        row["match_status"] = "매칭완료"
        append_daily(row, item)

    for row in rows:
        if row.get("retailer") == RETAILER:
            merge_daily(row)

    manual = update_manual_sales(items)
    summary = {
        "line_items": len(items),
        "payment": sum(item["payment"] for item in items),
        "qty": sum(item["qty"] for item in items),
        "orders": len(items),
        "created_rows": len(created),
        "created_preview": created[:40],
        "statuses": dict(collections.Counter(item["status"] for item in items)),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.dry_run:
        return
    DATA_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    MANUAL_FILE.write_text(json.dumps(manual, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
