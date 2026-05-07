import argparse
import collections
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl


DATA_FILE = Path("data.json")
MANUAL_FILE = Path("manual_sales_updates.json")
RETAILER = "자사몰"


def normalize_text(value):
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_name(name):
    text = normalize_text(name)
    text = re.sub(r"^\*", "", text).strip()
    text = text.replace("[채코제에디션]", "채코제에디션 ").strip()
    text = text.replace("퀵 드라이", "퀵드라이")
    text = text.replace("세비오버핏", "세미오버핏")
    text = text.replace("세미 오버핏", "세미오버핏")
    text = text.replace("러닝 팬츠", "러닝팬츠")
    text = text.replace("채코제 에디션", "채코제에디션")
    text = text.replace("차코제에디션", "채코제에디션")
    text = text.replace("채코제에디션퀵드라이 그린퓨처 티셔츠", "퀵드라이 그린퓨처 티셔츠")
    text = text.replace("채코제에디션 퀵드라이 그린퓨처 티셔츠", "퀵드라이 그린퓨처 티셔츠")
    text = text.replace("빈티지 세미크롭 스웻셔츠", "빈티지 워싱 스웻셔츠")
    text = text.replace("MESH ON", "MESH-ON")
    text = text.replace("MESH-ON 파워리프팅 롱티 (블랙)(하프 슬리브)", "MESH-ON 파워리프팅 롱티(하프 슬리브)")
    text = text.replace("[헤비쮸리] ", "")
    text = text.replace("[헤비쭈리] ", "")
    text = text.replace("빈티지 맨투맨 (세미오버핏)", "빈티지 맨투맨 (세미오버핏)")
    text = text.replace("빈티지 조거팬츠 (베이직)", "빈티지 조거팬츠 (베이직)")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_size(size):
    text = normalize_text(size).upper()
    return {"XXL": "2XL", "XXXL": "3XL"}.get(text, text)


def normalize_color(color):
    text = normalize_text(color).upper()
    aliases = {
        "BLACK": "블랙",
        "WHITE": "화이트",
        "CHARCOAL": "차콜",
        "DEEP CHARCOAL": "차콜",
        "GRAY": "그레이",
        "GREY": "그레이",
        "LIGHTGRAY": "라이트그레이",
        "LIGHT GRAY": "라이트그레이",
        "LIGHTGREY": "라이트그레이",
        "LIGHT GREY": "라이트그레이",
        "MELANGEGRAY": "멜란지그레이",
        "MELANGE GRAY": "멜란지그레이",
        "MELANGEGREY": "멜란지그레이",
        "MELANGE GREY": "멜란지그레이",
        "NAVY": "네이비",
        "BROWN": "브라운",
        "KHAKI GRAY": "카키그레이",
        "KHAKI GREY": "카키그레이",
        "VINTAGE KHAKI": "카키",
        "빈티지차콜": "차콜",
        "빈티지 카키": "카키",
        "딥차콜": "차콜",
        "차콜그레이": "차콜",
    }
    return aliases.get(text, normalize_text(color))


def parse_option(option):
    color = ""
    size = ""
    option = normalize_text(option)
    if not option:
        return color, size

    parts = re.split(r",\s*|\n+", option)
    for part in parts:
        if "=" in part:
            key, value = [x.strip() for x in part.split("=", 1)]
        elif ":" in part:
            key, value = [x.strip() for x in part.split(":", 1)]
        else:
            continue
        if key in {"색상", "컬러", "컬러값"}:
            color = value
        elif key in {"사이즈", "size", "SIZE"}:
            size = value
        elif key in {"디자인"} and not color:
            color = value
    return color, size


def as_int(value):
    if value in (None, ""):
        return 0
    return int(round(float(str(value).replace(",", ""))))


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def parse_order_key(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def allocate_order(rows):
    base_total = sum(as_int(row["상품구매금액(KRW)"]) for row in rows)
    final_payment = as_int(rows[0]["최종 결제금액(KRW)"])
    if base_total <= 0:
        payments = [0 for _ in rows]
    else:
        raw = [as_int(row["상품구매금액(KRW)"]) * final_payment / base_total for row in rows]
        payments = [int(round(v)) for v in raw]
        diff = final_payment - sum(payments)
        if payments:
            payments[-1] += diff
    return payments


def load_month_close(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    raw_rows = [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in values)
    ]

    grouped = collections.defaultdict(list)
    for row in raw_rows:
        day = parse_date(row["결제일시(입금확인일)"])
        if not day.startswith("2026-04-"):
            continue
        if row.get("환불완료일"):
            continue
        grouped[parse_order_key(row["결제일시(입금확인일)"])].append(row)

    items = []
    order_total_by_day = collections.Counter()
    order_count_by_day = collections.Counter()
    qty_by_day = collections.Counter()
    inconsistent_orders = []

    for order_key, order_rows in sorted(grouped.items()):
        final_values = {as_int(row["최종 결제금액(KRW)"]) for row in order_rows}
        if len(final_values) > 1:
            inconsistent_orders.append((order_key, sorted(final_values)))
        payments = allocate_order(order_rows)
        day = parse_date(order_rows[0]["결제일시(입금확인일)"])
        order_total_by_day[day] += as_int(order_rows[0]["최종 결제금액(KRW)"])
        order_count_by_day[day] += 1
        qty_by_day[day] += sum(as_int(row["수량"]) for row in order_rows)

        for row, payment in zip(order_rows, payments):
            color, size = parse_option(row.get("상품옵션(기본)"))
            name = normalize_name(row.get("상품명(한국어 쇼핑몰)"))
            color = normalize_color(color)
            size = normalize_size(size)
            qty = as_int(row.get("수량"))
            gross = as_int(row.get("상품구매금액(KRW)"))
            items.append(
                {
                    "date": day,
                    "order_key": order_key,
                    "name": name,
                    "color": color,
                    "size": size,
                    "qty": qty,
                    "gross": gross,
                    "payment": payment,
                }
            )

    return {
        "items": items,
        "daily": {
            day: {
                "payment": int(order_total_by_day[day]),
                "orders": int(order_count_by_day[day]),
                "qty": int(qty_by_day[day]),
            }
            for day in sorted(order_total_by_day)
        },
        "order_count": len(grouped),
        "inconsistent_orders": inconsistent_orders,
    }


def match_key(row):
    return (
        normalize_name(row.get("name", "")),
        normalize_color(row.get("color", "")),
        normalize_size(row.get("size", "")),
    )


def build_index(rows):
    index = collections.defaultdict(list)
    for i, row in enumerate(rows):
        if row.get("retailer") != RETAILER:
            continue
        index[match_key(row)].append(i)
    return index


def find_match(rows, index, item):
    keys = [
        (item["name"], item["color"], item["size"]),
        (item["name"], "", item["size"]),
        (item["name"], item["color"], ""),
        (item["name"], "", ""),
    ]
    for key in keys:
        hits = index.get(key)
        if hits:
            return hits[0]

    for i, row in enumerate(rows):
        if row.get("retailer") != RETAILER:
            continue
        row_name, row_color, row_size = match_key(row)
        if row_name != item["name"]:
            continue
        if item["color"] and row_color and item["color"] != row_color:
            continue
        if item["size"] and row_size and item["size"] != row_size:
            continue
        return i
    return None


def remove_existing_april_selfmall(rows):
    for row in rows:
        if row.get("retailer") != RETAILER:
            continue
        kept_daily = []
        removed_qty = removed_gross = removed_payment = removed_orders = 0
        for daily in row.get("daily", []):
            if str(daily.get("date", "")).startswith("2026-04-"):
                removed_qty += as_int(daily.get("qty"))
                removed_gross += as_int(daily.get("gross"))
                removed_payment += as_int(daily.get("payment"))
                removed_orders += as_int(daily.get("orders"))
            else:
                kept_daily.append(daily)
        if len(kept_daily) != len(row.get("daily", [])):
            row["daily"] = kept_daily
            row["qty"] = as_int(row.get("qty")) - removed_qty
            row["gross"] = as_int(row.get("gross")) - removed_gross
            row["payment"] = as_int(row.get("payment")) - removed_payment
            row["orders"] = as_int(row.get("orders")) - removed_orders
            row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def append_daily(row, item):
    row.setdefault("daily", []).append(
        {
            "date": item["date"],
            "qty": item["qty"],
            "gross": item["gross"],
            "payment": item["payment"],
            "orders": 1,
        }
    )
    row["qty"] = as_int(row.get("qty")) + item["qty"]
    row["gross"] = as_int(row.get("gross")) + item["gross"]
    row["payment"] = as_int(row.get("payment")) + item["payment"]
    row["orders"] = as_int(row.get("orders")) + 1
    row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def merge_same_day_daily(rows):
    for row in rows:
        merged = {}
        for daily in row.get("daily", []):
            day = daily.get("date")
            current = merged.setdefault(day, {"date": day, "qty": 0, "gross": 0, "payment": 0, "orders": 0})
            current["qty"] += as_int(daily.get("qty"))
            current["gross"] += as_int(daily.get("gross"))
            current["payment"] += as_int(daily.get("payment"))
            current["orders"] += as_int(daily.get("orders"))
        row["daily"] = [merged[day] for day in sorted(merged)]


def update_manual_sales(daily):
    manual = json.loads(MANUAL_FILE.read_text(encoding="utf-8"))
    by_date = {entry["date"]: entry for entry in manual}
    for day, totals in daily.items():
        entry = by_date.setdefault(day, {"date": day, "retailers": []})
        retailers = entry.setdefault("retailers", [])
        target = next((r for r in retailers if r.get("retailer") == RETAILER), None)
        if target is None:
            target = {"retailer": RETAILER, "items": []}
            retailers.append(target)
        target["payment"] = totals["payment"]
        target["orders"] = totals["orders"]
        target["qty"] = totals["qty"]
        target["items"] = []
    return [by_date[day] for day in sorted(by_date)]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    close = load_month_close(Path(args.xlsx))
    rows = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    remove_existing_april_selfmall(rows)
    index = build_index(rows)
    unmatched = []

    for item in close["items"]:
        match_idx = find_match(rows, index, item)
        if match_idx is None:
            unmatched.append(item)
            new_row = {
                "retailer": RETAILER,
                "mall_no": "",
                "name": item["name"],
                "color": item["color"],
                "size": item["size"],
                "qty": 0,
                "gross": 0,
                "payment": 0,
                "avg_unit": 0,
                "orders": 0,
                "source_type": "상품",
                "daily": [],
                "match_status": "미매칭",
                "match_sku": "",
                "standard_name": f"{item['name']}_{item['color']}-{item['size']}".strip("_-"),
                "received_qty": 0,
                "stock_qty": 0,
            }
            rows.append(new_row)
            index[(item["name"], item["color"], item["size"])].append(len(rows) - 1)
            match_idx = len(rows) - 1
        append_daily(rows[match_idx], item)

    merge_same_day_daily(rows)
    manual = update_manual_sales(close["daily"])

    summary = {
        "line_items": len(close["items"]),
        "orders": close["order_count"],
        "payment": sum(day["payment"] for day in close["daily"].values()),
        "qty": sum(day["qty"] for day in close["daily"].values()),
        "unmatched_count": len(unmatched),
        "inconsistent_orders": close["inconsistent_orders"],
        "daily": close["daily"],
        "unmatched_preview": unmatched[:30],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.dry_run:
        return

    DATA_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    MANUAL_FILE.write_text(json.dumps(manual, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
