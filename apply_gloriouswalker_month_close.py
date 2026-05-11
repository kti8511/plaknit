import argparse
import collections
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl


DATA_FILE = Path("data.json")
MANUAL_FILE = Path("manual_sales_updates.json")
RETAILER = "글로리어스워커"

NAME_MAP = {
    "Essential Heat Running Pants": ("ESSENTIAL HEAT 기모 러닝팬츠", ""),
    "Quickdry Ventilation Running Top": ("퀵드라이 벤틸레이션 러닝탑", ""),
    "Quickdry Two-in-One Running Shorts": ("퀵드라이 투인원 러닝쇼츠", ""),
}

COLOR_MAP = {
    "BLACK": "블랙",
    "BROWN": "브라운",
    "GREY": "그레이",
    "GRAY": "그레이",
}


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def as_int(value):
    if value in (None, ""):
        return 0
    return int(round(float(str(value).replace(",", ""))))


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def parse_size(option):
    text = clean_text(option)
    match = re.search(r"사이즈\s*=\s*([^()]+)", text)
    return clean_text(match.group(1) if match else text)


def parse_product(raw_name):
    raw = clean_text(raw_name)
    raw = re.sub(r"^\[[^\]]+\]\s*", "", raw).strip()
    if "_" in raw:
        name_part, color_part = raw.rsplit("_", 1)
    else:
        name_part, color_part = raw, ""
    name = NAME_MAP.get(name_part, (name_part, ""))[0]
    color = COLOR_MAP.get(color_part.upper(), color_part)
    return name, color


def standard_name(item):
    if item["color"] and item["size"]:
        return f"{item['name']}_{item['color']}-{item['size']}"
    if item["color"]:
        return f"{item['name']}_{item['color']}"
    if item["size"]:
        return f"{item['name']}-{item['size']}"
    return item["name"]


def load_items(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    items = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        row = dict(zip(headers, values))
        qty = as_int(row.get("수량"))
        gross = as_int(row.get("상품구매금액"))
        payment = as_int(row.get("총 결제금액"))
        if qty == 0 and gross == 0 and payment == 0:
            continue
        name, color = parse_product(row.get("주문상품명"))
        size = parse_size(row.get("상품옵션"))
        items.append(
            {
                "date": parse_date(row.get("주문일시")),
                "name": name,
                "color": color,
                "size": size,
                "qty": qty,
                "gross": gross,
                "payment": payment,
            }
        )
    return items


def row_key(row):
    return (
        clean_text(row.get("name")),
        clean_text(row.get("color")),
        clean_text(row.get("size")),
    )


def remove_month(rows, month):
    for row in rows:
        if row.get("retailer") != RETAILER:
            continue
        kept = []
        removed_qty = removed_gross = removed_payment = removed_orders = 0
        for daily in row.get("daily", []):
            if str(daily.get("date", "")).startswith(month):
                removed_qty += as_int(daily.get("qty"))
                removed_gross += as_int(daily.get("gross"))
                removed_payment += as_int(daily.get("payment"))
                removed_orders += as_int(daily.get("orders"))
            else:
                kept.append(daily)
        row["daily"] = kept
        row["qty"] = as_int(row.get("qty")) - removed_qty
        row["gross"] = as_int(row.get("gross")) - removed_gross
        row["payment"] = as_int(row.get("payment")) - removed_payment
        row["orders"] = as_int(row.get("orders")) - removed_orders
        row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def build_index(rows):
    index = {}
    standard_index = {}
    for idx, row in enumerate(rows):
        if row.get("retailer") != RETAILER:
            continue
        index.setdefault(row_key(row), idx)
        if row.get("standard_name"):
            standard_index.setdefault(row["standard_name"], idx)
    return index, standard_index


def append_daily(row, item):
    for daily in row.setdefault("daily", []):
        if daily.get("date") == item["date"]:
            daily["qty"] = as_int(daily.get("qty")) + item["qty"]
            daily["gross"] = as_int(daily.get("gross")) + item["gross"]
            daily["payment"] = as_int(daily.get("payment")) + item["payment"]
            daily["orders"] = as_int(daily.get("orders")) + max(item["qty"], 0)
            break
    else:
        row["daily"].append(
            {
                "date": item["date"],
                "qty": item["qty"],
                "gross": item["gross"],
                "payment": item["payment"],
                "orders": max(item["qty"], 0),
            }
        )
    row["qty"] = as_int(row.get("qty")) + item["qty"]
    row["gross"] = as_int(row.get("gross")) + item["gross"]
    row["payment"] = as_int(row.get("payment")) + item["payment"]
    row["orders"] = as_int(row.get("orders")) + max(item["qty"], 0)
    row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def update_manual(items, month):
    manual = json.loads(MANUAL_FILE.read_text(encoding="utf-8"))
    for entry in manual:
        if str(entry.get("date", "")).startswith(month):
            entry["retailers"] = [
                retailer
                for retailer in entry.get("retailers", [])
                if retailer.get("retailer") != RETAILER
            ]

    grouped = collections.defaultdict(list)
    for item in items:
        grouped[item["date"]].append(item)
    by_date = {entry["date"]: entry for entry in manual}
    for day, day_items in grouped.items():
        entry = by_date.setdefault(day, {"date": day, "retailers": []})
        target = {
            "retailer": RETAILER,
            "payment": sum(item["payment"] for item in day_items),
            "qty": sum(item["qty"] for item in day_items),
            "orders": sum(max(item["qty"], 0) for item in day_items),
            "items": [
                {"name": standard_name(item), "qty": item["qty"], "payment": item["payment"]}
                for item in day_items
            ],
        }
        entry.setdefault("retailers", []).append(target)
    return [by_date[date] for date in sorted(by_date)]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("--month", required=True, help="YYYY-MM")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    items = load_items(Path(args.xlsx))
    rows = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    remove_month(rows, args.month)
    index, standard_index = build_index(rows)
    created = []

    for item in items:
        std = standard_name(item)
        idx = standard_index.get(std)
        if idx is None:
            idx = index.get((item["name"], item["color"], item["size"]))
        if idx is None:
            new_row = {
                "retailer": RETAILER,
                "mall_no": "",
                "name": item["name"],
                "color": item["color"],
                "size": item["size"],
                "qty": 0,
                "gross": 0,
                "payment": 0,
                "avg_unit": 0,
                "orders": 0,
                "source_type": "단품",
                "daily": [],
                "match_status": "매칭완료",
                "match_sku": "",
                "standard_name": std,
                "received_qty": 0,
                "stock_qty": 0,
            }
            rows.append(new_row)
            idx = len(rows) - 1
            index[(item["name"], item["color"], item["size"])] = idx
            standard_index[std] = idx
            created.append(std)
        row = rows[idx]
        row["name"] = item["name"]
        row["color"] = item["color"]
        row["size"] = item["size"]
        row["standard_name"] = std
        row["match_status"] = "매칭완료"
        append_daily(row, item)

    manual = update_manual(items, args.month)
    summary = {
        "month": args.month,
        "days": sorted({item["date"] for item in items}),
        "line_items": len(items),
        "qty": sum(item["qty"] for item in items),
        "gross": sum(item["gross"] for item in items),
        "payment": sum(item["payment"] for item in items),
        "created_rows": created,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if args.dry_run:
        return
    DATA_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    MANUAL_FILE.write_text(json.dumps(manual, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
