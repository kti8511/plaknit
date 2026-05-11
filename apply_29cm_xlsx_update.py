import argparse
import collections
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl


DATA_FILE = Path("data.json")
MANUAL_FILE = Path("manual_sales_updates.json")
RETAILER = "29cm"


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def as_int(value):
    if value in (None, ""):
        return 0
    return int(round(float(str(value).replace(",", ""))))


def norm_size(value):
    text = clean_text(value).upper()
    return {"XXL": "2XL", "XXXL": "3XL", "XXXXL": "4XL"}.get(text, text)


def norm_color(value):
    text = clean_text(value)
    return {
        "BLACK": "블랙",
        "WHITE": "화이트",
        "CHARCOAL": "차콜",
    }.get(text.upper(), text)


def normalize_name(name):
    text = clean_text(name)
    text = text.replace("퀵 드라이", "퀵드라이")
    text = text.replace("세미 오버핏", "세미오버핏")
    return text


def split_name_color(name, color):
    name = normalize_name(name)
    color = norm_color(color)
    if color:
        return name, color
    suffixes = [
        (" 라이트그레이", "라이트그레이"),
        (" 멜란지그레이", "멜란지그레이"),
        (" 카키그레이", "카키그레이"),
        (" 블랙", "블랙"),
        (" 화이트", "화이트"),
        (" 차콜", "차콜"),
        (" 그레이", "그레이"),
    ]
    for suffix, suffix_color in suffixes:
        if name.endswith(suffix):
            return name[: -len(suffix)].strip(), suffix_color
    return name, color


def parse_option(option):
    color = ""
    size = ""
    option = clean_text(option)
    for key, value in re.findall(r"\[(색상|사이즈)\]\s*([^,\]]+)", option):
        if key == "색상":
            color = norm_color(value)
        elif key == "사이즈":
            size = norm_size(value)
    return color, size


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def standard_name(item):
    if item["color"] and item["size"]:
        return f"{item['name']}_{item['color']}-{item['size']}"
    if item["color"]:
        return f"{item['name']}_{item['color']}"
    if item["size"]:
        return f"{item['name']}-{item['size']}"
    return item["name"]


def load_items(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    items = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        row = dict(zip(headers, values))
        day = parse_date(row.get("주문일시"))
        color, size = parse_option(row.get("옵션"))
        name, color = split_name_color(row.get("상품명"), color)
        qty = as_int(row.get("수량"))
        gross = as_int(row.get("판매액"))
        payment = as_int(row.get("실 판매액"))
        if qty == 0 and gross == 0 and payment == 0:
            continue
        items.append(
            {
                "date": day,
                "name": name,
                "color": color,
                "size": size,
                "qty": qty,
                "gross": gross,
                "payment": payment,
            }
        )
    return items


def row_key(row):
    return (
        normalize_name(row.get("name")),
        norm_color(row.get("color")),
        norm_size(row.get("size")),
    )


def remove_existing_days(rows, days, replace_month=None):
    for row in rows:
        if row.get("retailer") != RETAILER:
            continue
        kept = []
        removed_qty = removed_gross = removed_payment = removed_orders = 0
        for daily in row.get("daily", []):
            should_remove = daily.get("date") in days
            if replace_month:
                should_remove = should_remove or str(daily.get("date", "")).startswith(replace_month)
            if should_remove:
                removed_qty += as_int(daily.get("qty"))
                removed_gross += as_int(daily.get("gross"))
                removed_payment += as_int(daily.get("payment"))
                removed_orders += as_int(daily.get("orders"))
            else:
                kept.append(daily)
        if len(kept) != len(row.get("daily", [])):
            row["daily"] = kept
            row["qty"] = as_int(row.get("qty")) - removed_qty
            row["gross"] = as_int(row.get("gross")) - removed_gross
            row["payment"] = as_int(row.get("payment")) - removed_payment
            row["orders"] = as_int(row.get("orders")) - removed_orders
            row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def build_index(rows):
    index = {}
    by_name = {}
    for idx, row in enumerate(rows):
        if row.get("retailer") != RETAILER:
            continue
        key = row_key(row)
        index.setdefault(key, idx)
        by_name.setdefault(key[0], idx)
    return index, by_name


def append_daily(row, item):
    row.setdefault("daily", []).append(
        {
            "date": item["date"],
            "qty": item["qty"],
            "gross": item["gross"],
            "payment": item["payment"],
            "orders": 1,
        }
    )
    row["qty"] = as_int(row.get("qty")) + item["qty"]
    row["gross"] = as_int(row.get("gross")) + item["gross"]
    row["payment"] = as_int(row.get("payment")) + item["payment"]
    row["orders"] = as_int(row.get("orders")) + 1
    row["avg_unit"] = int(round(row["payment"] / row["qty"])) if row["qty"] else 0


def update_manual(items, replace_month=None):
    manual = json.loads(MANUAL_FILE.read_text(encoding="utf-8"))
    if replace_month:
        for entry in manual:
            if not str(entry.get("date", "")).startswith(replace_month):
                continue
            entry["retailers"] = [
                retailer
                for retailer in entry.get("retailers", [])
                if retailer.get("retailer") != RETAILER
            ]
    grouped = collections.defaultdict(list)
    for item in items:
        grouped[item["date"]].append(item)
    by_date = {entry["date"]: entry for entry in manual}
    for day, day_items in grouped.items():
        entry = by_date.setdefault(day, {"date": day, "retailers": []})
        retailers = entry.setdefault("retailers", [])
        target = next((retailer for retailer in retailers if retailer.get("retailer") == RETAILER), None)
        if target is None:
            target = {"retailer": RETAILER, "items": []}
            retailers.append(target)
        target["payment"] = sum(item["payment"] for item in day_items)
        target["qty"] = sum(item["qty"] for item in day_items)
        target["orders"] = len(day_items)
        target["items"] = [
            {"name": standard_name(item), "qty": item["qty"], "payment": item["payment"]}
            for item in day_items
        ]
    return [by_date[date] for date in sorted(by_date)]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("--replace-month", help="YYYY-MM 형태로 해당 월의 29CM 데이터를 전체 교체")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    items = load_items(Path(args.xlsx))
    rows = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    days = {item["date"] for item in items}
    remove_existing_days(rows, days, args.replace_month)
    index, by_name = build_index(rows)
    created = []
    for item in items:
        key = (item["name"], item["color"], item["size"])
        idx = index.get(key) or by_name.get(item["name"])
        if idx is None:
            new_row = {
                "retailer": RETAILER,
                "mall_no": "",
                "name": item["name"],
                "color": item["color"],
                "size": item["size"],
                "qty": 0,
                "gross": 0,
                "payment": 0,
                "avg_unit": 0,
                "orders": 0,
                "source_type": "상품",
                "daily": [],
                "match_status": "매칭완료",
                "match_sku": "",
                "standard_name": standard_name(item),
                "received_qty": 0,
                "stock_qty": 0,
            }
            rows.append(new_row)
            idx = len(rows) - 1
            index[key] = idx
            by_name[item["name"]] = idx
            created.append(new_row["standard_name"])
        row = rows[idx]
        row["name"] = item["name"]
        row["color"] = item["color"]
        row["size"] = item["size"]
        row["standard_name"] = standard_name(item)
        row["match_status"] = "매칭완료"
        append_daily(row, item)

    manual = update_manual(items, args.replace_month)
    summary = {
        "days": sorted(days),
        "line_items": len(items),
        "qty": sum(item["qty"] for item in items),
        "payment": sum(item["payment"] for item in items),
        "created_rows": created,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if args.dry_run:
        return
    DATA_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    MANUAL_FILE.write_text(json.dumps(manual, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
